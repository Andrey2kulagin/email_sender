import openpyxl
from rest_framework.response import Response
from django.core.files.uploadedfile import InMemoryUploadedFile
import os
from ..models import User, RecipientContact, ContactGroup, ContactImportFiles
from .all_service import check_or_create_folder, phone_normalize
from rest_framework import serializers, exceptions
from django.core.exceptions import ObjectDoesNotExist
import copy
from .contact_service import email_validate, phone_validate


def gen_path_to_import_report(user, import_id):
    try:
        ContactImportFiles.objects.get(owner=user, id=import_id)
        return f"sources/contact_bug_reports/{user.username}/errors_import{import_id}.xlsx"
    except ObjectDoesNotExist:
        return 404


def contact_import(validate_data: dict, user: User):
    filename = validate_data.get("filename")
    statistics = {
        "OK": 0,
        "PF": 0,
        "FF": 0,
        "all_count": 0
    }
    if filename:
        mask = copy.deepcopy(validate_data)
        del mask["filename"]
        error_rows = [get_errors_headers(mask)]
        workbook = openpyxl.load_workbook(f"sources/import_file/{user}/{filename}")
        sheet = workbook.active
        import_obj = ContactImportFiles.objects.get(owner=user, filename=filename)
        is_contains_headers_flag = import_obj.is_contains_headers
        # обработка самого файла построчно
        for row in sheet.iter_rows(values_only=True):
            # валидация строки запись контакта в БД
            if is_contains_headers_flag:
                is_contains_headers_flag = False
                continue
            statistics[import_contact_row_handler(row, mask, user, error_rows)] += 1
            statistics["all_count"] += 1
        # запись данных об ошибках в файл
        import_obj.is_imported = True
        import_obj.all_handled_lines = statistics["all_count"]
        import_obj.success_create_update_count = statistics["OK"]
        import_obj.partial_success_create_update_count = statistics["PF"]
        import_obj.fail_count = statistics["FF"]
        import_obj.save()
        write_errors(error_rows, user, import_obj.id)

    return statistics


def write_array(filename, array):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in array:
        ws.append(row)
    wb.save(filename)


def write_errors(error_rows, user, import_id):
    dir_path = f"sources/contact_bug_reports/{user.username}/"
    check_or_create_folder(dir_path)
    filename = f"errors_import{import_id}.xlsx"
    write_array(dir_path + filename, error_rows)


def get_errors_headers(mask):
    row = ["Критическая ошибка(создать не удалось)", "Некритическая ошибка (удалось создать частично)"]
    if mask.get("id") is None:
        row.append("id")
    keys_row = ["" for i in range(23)]
    for key in mask:
        cure_index = mask[key]
        if type(cure_index) == int:
            keys_row[cure_index] = key
    row.extend(keys_row)
    return row


def import_contact_row_handler(row: list[str], mask: dict[str: int], user: User, error_rows: list[list[str]]):
    cure_values = get_cure_value(row, mask)
    contact_id = cure_values.get("id")
    validate_result = errors_check(contact_id, cure_values, user)
    status = validate_result.get("status", "")
    errors = validate_result.get("errors")
    comment = validate_result.get("comment")
    if status in ["OK", "PF"]:
        contact_id = save_contact(user, cure_values, errors)
        if contact_id and status == "PF":
            error_rows.append(get_error_row(errors, row, comment, contact_id, mask["id"], status))
    elif status == "FF":
        error_rows.append(get_error_row(errors, row, comment, None, mask["id"], status))
    return status


def errors_check(contact_id: str, cure_values: dict, user: User):
    # если есть id, то обновляем, если нет, то создаём
    if contact_id:
        validate_result = import_contact_update_validate_errors_check(cure_values, contact_id, user)
    else:
        validate_result = import_contact_create_validate_error_check(cure_values, user)
    return validate_result


def get_error_row(errors: dict, row: list[str], comment: str, contact_id: (int, None), id_index: int, status) -> list[
    str]:
    """
    print("ERRORS", errors, "\n")
    print("ROW", row, "\n")
    print("COMMENT", comment, "\n")
    print("CONTACT_ID", contact_id, "\n")
    print("ID_INDEX", id_index, "\n")
    print("STATUS", status, "\n\n\n\n\n\n\n")"""

    error_row = []
    error_str = ""
    row = list(row)
    # добавляем коммент
    if comment:
        error_str += comment
        error_str += "\n\n"
    error_str += "errors:\n"
    count = 1
    # добавляем ошибки
    for error in errors:
        error_str += f"{count}. {errors[error]}\n"
    if status == "PF":
        error_row.append("")
        error_row.append(error_str)
    else:
        error_row.append(error_str)
        error_row.append("")
    # добавляем id
    if id_index is not None:
        if row[id_index] is None:
            if contact_id:
                row[id_index] = str(contact_id)
            else:
                row[id_index] = ""
    else:
        if contact_id:
            error_row.append(str(contact_id))
        else:
            error_row.append("")
    error_row.extend(row)
    return error_row


def save_contact(user: User, cure_values: dict, errors: dict):
    cure_contact_id = cure_values.get("id")
    contact = get_cure_contact_obj_to_save(cure_contact_id, user)
    if contact is None:
        return
    contact_save_group_handler(cure_values, user, contact)

    contact_save_another_params_handler(cure_values, errors, contact)
    contact.save()
    return contact.id


def contact_save_another_params_handler(cure_values: dict, errors, contact):
    for key in cure_values:
        value = cure_values.get(key)
        if value and ((errors and key not in errors) or not errors):
            setattr(contact, key, value)


def contact_save_group_handler(cure_values: dict, user: User, contact: RecipientContact):
    groups = cure_values.get("contact_group")

    if groups:

        groups = groups.split(";")
        for group in groups:
            try:
                group = ContactGroup.objects.get(user=user, title=group.strip())
                contact.contact_group.add(group)
            except ObjectDoesNotExist:
                continue
    del cure_values["contact_group"]


def get_cure_contact_obj_to_save(cure_contact_id, user):
    if cure_contact_id:
        try:
            contact = RecipientContact.objects.get(owner=user, id=cure_contact_id)
        except ObjectDoesNotExist:
            return
    else:
        contact = RecipientContact()
        contact.owner = user
        contact.save()
    return contact


def import_contact_update_validate_errors_check(cure_values, contact_id, user):
    # здесь будем возвращать строку из 3 статусов - OK, FF(full fail), PF(partial fail)
    # если что-то хотя бы 1 параметр валиден, а остальные - нет, то возвращаем PF(обновляем только то, что валидно)
    errors = {}
    contact_objs = RecipientContact.objects.filter(owner=user, id=contact_id)
    if len(contact_objs) == 0:
        errors["id"] = "У вас нет контакта с таким id"
        return {"status": "FF", "errors": errors, "comment": None}
    add_email_phone_errors(cure_values, user, errors)
    add_groups_errors(cure_values, user, errors)
    if len(errors) == 0:
        return {"status": "OK", "errors": None, "comment": None}
    return {"status": "PF", "errors": errors, "comment": None}


def add_email_phone_errors(cure_values, user, errors):
    try:
        cure_phone = cure_values["phone"]
        if cure_phone:
            cure_values["phone"] = phone_normalize(cure_phone)
        phone_validate(cure_values, "Неправильный номер телефона. Должно быть 11 цифр и начинаться должен с 8 или +7",
                       user)
    except serializers.ValidationError as e:
        errors["phone"] = e.__str__()
    try:
        email_validate(cure_values, "Введите правильный email", user)
    except serializers.ValidationError as e:
        errors["email"] = e.__str__()


def import_contact_create_validate_error_check(cure_values, user):
    # здесь будем возвращать строку из 3 статусов - OK, FF(full fail), PF(partial fail)
    # Классическая ситуация с валидацией. Должен быть хотя бы 1 рабочий контакт
    errors = {}
    add_email_phone_errors(cure_values, user, errors)
    input_phone = cure_values.get("phone")
    phone_errors = errors.get("phone")
    input_email = cure_values.get("email")
    email_errors = errors.get("email")
    if input_phone is None and email_errors:
        return {"status": "FF", "errors": errors, "comment": "Добавьте хотя бы 1 валидный контакт", }
    if input_email is None and phone_errors:
        return {"status": "FF", "errors": errors, "comment": "Добавьте хотя бы 1 валидный контакт", }
    if phone_errors and email_errors:
        return {"status": "FF", "errors": errors, "comment": "Добавьте хотя бы 1 валидный контакт", }
    add_groups_errors(cure_values, user, errors)
    if len(errors) == 0:
        return {"status": "OK", "errors": None, "comment": None}
    else:
        return {"status": "PF", "errors": errors, "comment": None}


def add_groups_errors(cure_values, user, errors):
    groups_str = cure_values.get("contact_group")
    if groups_str and groups_str != "DELETE_ALL":
        groups = groups_str.split(";")
        for group in groups:
            if group:
                try:
                    ContactGroup.objects.get(user=user, title=group.strip())
                except ObjectDoesNotExist:
                    errors[f"group :{group}"] = f'У вас нет группы: "{group.strip()}"'


def get_cure_value(row: list[str], mask: dict[str: int]):
    cure_values = {}
    for key in mask:
        index = mask[key]
        try:
            value = row[index]
            if value is not None:
                cure_values[key] = str(value)
            else:
                cure_values[key] = None
        except (IndexError, TypeError):
            cure_values[key] = None
    return cure_values


def contact_import_run_request_data_validate(data, user):
    # прописать валидацию номеров ячеек
    filename = data.get("filename")
    try:
        obj = ContactImportFiles.objects.get(owner=user, filename=filename)
        if obj.is_imported:
            raise serializers.ValidationError("Этот файл уже импортирован")
        row_len = obj.row_len
        param_fields_name = ["id", "name", "surname", "email", "contact_group", "comment"]
        fail_params = []
        is_have_fail_params = False
        for param in param_fields_name:
            request_param = data.get(param)
            if request_param and request_param >= row_len:
                fail_params.append(param)
                is_have_fail_params = True
        if is_have_fail_params:
            raise serializers.ValidationError(
                f"Переданы индексы, которые больше чем длина строки в файле. Поля с ошибкой:{fail_params}. Должны быть индексы до {row_len}")
    except ObjectDoesNotExist:
        raise exceptions.NotFound("Такого файла нет на сервере")
    if not (data.get("phone") or data.get("email")):
        raise serializers.ValidationError("Должен быть хотя бы 1 контакт")


def file_upload_handler(file: InMemoryUploadedFile, user: User, is_contains_headers: bool):
    start_data = get_start_data_from_imp(file)
    file_dir = f'sources/import_file/{user.username}'
    filename = get_cure_filename(file_dir, file.name)
    check_or_create_folder(file_dir)
    write_received_excel(file_dir, filename, file)
    contact_import_obj = ContactImportFiles.objects.create(owner=user, filename=filename,
                                                           row_len=start_data.get("count_elements_in_line"),
                                                           is_contains_headers=is_contains_headers)
    output_data = {
        "import_id": contact_import_obj.id,
        "file_name_on_serv": filename,
        "count": start_data.get("count"),
        "count_elements_in_line": start_data.get("count_elements_in_line"),
        "results": start_data.get("results")
    }
    return Response(data=output_data, status=200)


def write_received_excel(file_dir, filename, file):
    with open(f"{file_dir}/{filename}", 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)


def get_cure_filename(filepath, init_filename):
    count = 1
    while os.path.exists(f"{filepath}/{init_filename}"):
        init_filename = init_filename.replace(".", f"{count}.")
        count += 1
    return init_filename


def is_file_in_dir(filepath):
    return os.path.exists(filepath)


def get_start_data_from_imp(file: InMemoryUploadedFile):
    workbook = openpyxl.load_workbook(file)
    # Получаем название всех листов в файле
    sheet_names = workbook.sheetnames
    # Выбираем первый лист
    sheet = workbook[sheet_names[0]]
    results = []
    sheet_len = 0
    max_item_length = 20
    for row in sheet:
        row_len = len(row)
        if row_len < max_item_length:
            max_item_length = row_len
        if sheet_len >= 10:
            break
        sheet_len += 1
        results.append(get_new_row(row[:20]))
    return {
        "count": sheet_len,
        "count_elements_in_line": max_item_length,
        "results": results
    }


def get_new_row(row):
    result_row = []
    for i in row:
        if i.value:
            result_row.append(str(i.value))
        else:
            result_row.append("")
    return result_row
